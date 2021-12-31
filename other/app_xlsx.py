import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row + 1):
        #print(row)
        cell = sheet.cell(row, 3)
        #print(cell.value)
        corrected_cell = cell.value * 0.9
        corrected_cell_price = sheet.cell(row, 4)
        corrected_cell_price.value = corrected_cell

#generate a chart 
    values = Reference(sheet, 
              min_row=2, 
              max_row=sheet.max_row, 
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'f2')

    wb.save(f'2{filename}')


process_workbook('transactions.xlsx')
